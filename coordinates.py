import requests
from pyproj import Proj, transform
from openpyxl import load_workbook

def get_lat_lon(address, api_key):
    """ Get latitude and longitude from address using Google's Geocoding API """
    base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {
        "address": address,
        "key": api_key
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        result = response.json()
        if result["status"] == "OK":
            location = result["results"][0]["geometry"]["location"]
            return location["lat"], location["lng"]
    return None, None

def convert_to_utm(lat, lon):
    """ Convert latitude and longitude to UTM coordinates """
    if not lat or not lon:
        return None, None
    proj_latlon = Proj(proj='latlong', datum='WGS84')
    proj_utm = Proj(proj='utm', zone=compute_utm_zone(lon), datum='WGS84')
    utm_x, utm_y = transform(proj_latlon, proj_utm, lon, lat)
    return utm_x, utm_y

def compute_utm_zone(lon):
    """ Compute UTM zone from longitude """
    return int((lon + 180) / 6) + 1

def process_addresses(excel_path, api_key, sheet_name='Output', max_addresses=10):
    """ Process each address in the Excel file, up to a maximum of max_addresses """
    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]

    address_count = 0
    for row in range(2, sheet.max_row + 1):
        if address_count >= max_addresses:
            break

        address = sheet.cell(row, 1).value
        if address:
            lat, lon = get_lat_lon(address, api_key)
            if lat and lon:
                # Combine lat and lon into one string separated by a comma
                lat_lon_combined = f"{lat}, {lon}"
                sheet.cell(row, 2).value = lat_lon_combined

                utm_x, utm_y = convert_to_utm(lat, lon)
                if utm_x and utm_y:
                    # Optionally, combine UTM coordinates into one string
                    utm_combined = f"{utm_x}, {utm_y}"
                    sheet.cell(row, 3).value = utm_combined
                else:
                    print(f"Could not convert to UTM coordinates for address: {address}")
            else:
                print(f"Could not get geocode for the address: {address}")
            
            address_count += 1

    wb.save(excel_path)

if __name__ == "__main__":
    excel_file_path = "Your File.xlsx" # Replace with File Name of input file
    google_api_key = "ApiKey"  # Replace with your actual API key
    process_addresses(excel_file_path, google_api_key)
